import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from db import get_db, db_cursor
from datetime import datetime

def generate_school_excel(school_id):
    db = get_db()
    
    # 1. Fetch School Data
    with db_cursor(db) as cursor:
        cursor.execute('SELECT * FROM schools WHERE id = %s', (school_id,))
        school = cursor.fetchone()
        if not school:
            return None

        # 2. Fetch Student Data
        cursor.execute('''
            SELECT u.username, sd.full_name, sd.email, sd.mobile, sd.admission_number, c.name as classroom
            FROM users u
            JOIN student_details sd ON u.id = sd.user_id
            LEFT JOIN classrooms c ON sd.classroom_id = c.id
            WHERE u.school_id = %s AND u.role = 'student'
        ''', (school_id,))
        students = cursor.fetchall()

        # 3. Fetch Teacher Data
        cursor.execute('''
            SELECT u.username, td.full_name, td.email, td.mobile, td.department
            FROM users u
            JOIN teacher_details td ON u.id = td.user_id
            WHERE u.school_id = %s AND u.role = 'teacher'
        ''', (school_id,))
        teachers = cursor.fetchall()

        # 4. Calculate Stats
        cursor.execute('SELECT COUNT(*) as count FROM users WHERE school_id = %s', (school_id,))
        total_users = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM courses WHERE school_id = %s', (school_id,))
        total_courses = cursor.fetchone()['count']

    # Create Workbook
    wb = Workbook()
    
    # --- Sheet 1: Overview ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    ws_overview.append(["Metric", "Value"])
    for cell in ws_overview[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    ws_overview.append(["Institution Name", school['name']])
    ws_overview.append(["Slug", school['slug']])
    ws_overview.append(["Total Users", total_users])
    ws_overview.append(["Total Students", len(students)])
    ws_overview.append(["Total Teachers", len(teachers)])
    ws_overview.append(["Total Courses", total_courses])
    ws_overview.append(["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    # Set column widths
    ws_overview.column_dimensions['A'].width = 25
    ws_overview.column_dimensions['B'].width = 40

    # --- Sheet 2: Students ---
    ws_students = wb.create_sheet(title="Students")
    student_headers = ["Admission #", "Full Name", "Username", "Email", "Mobile", "Classroom"]
    ws_students.append(student_headers)
    for cell in ws_students[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for s in students:
        ws_students.append([s['admission_number'], s['full_name'], s['username'], s['email'], s['mobile'], s['classroom']])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_students.column_dimensions[col].width = 20

    # --- Sheet 3: Teachers ---
    ws_teachers = wb.create_sheet(title="Teachers")
    teacher_headers = ["Full Name", "Username", "Email", "Mobile", "Department"]
    ws_teachers.append(teacher_headers)
    for cell in ws_teachers[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for t in teachers:
        ws_teachers.append([t['full_name'], t['username'], t['email'], t['mobile'], t['department']])
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_teachers.column_dimensions[col].width = 20

    # Save to temp file
    filename = f"school_report_{school['slug']}_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    temp_path = os.path.join("uploads", filename)
    if not os.path.exists("uploads"):
        os.makedirs("uploads")
        
    wb.save(temp_path)
    return temp_path
