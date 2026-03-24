from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from db import get_db, db_cursor
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

classrooms_bp = Blueprint('classrooms', __name__)


def _require_staff():
    if current_user.role not in ['teacher', 'admin', 'principal']:
        flash('Unauthorized access.', 'error')
        return redirect(url_for('dashboard.dashboard'))
    return None


# ── List all classrooms ─────────────────────────────────────────────────────

@classrooms_bp.route('/classrooms')
@login_required
def index():
    if current_user.role == 'student':
        db = get_db()
        with db_cursor(db) as cursor:
            cursor.execute('SELECT classroom_id FROM student_details WHERE user_id = %s AND school_id = %s', (current_user.id, current_user.school_id))
            row = cursor.fetchone()
            if row and row['classroom_id']:
                return redirect(url_for('classrooms.detail', classroom_id=row['classroom_id']))
            else:
                flash('You are not assigned to a classroom yet.', 'info')
                return redirect(url_for('dashboard.dashboard'))

    redir = _require_staff()
    if redir:
        return redir

    db = get_db()
    with db_cursor(db) as cursor:
        if current_user.role in ['admin', 'principal']:
            cursor.execute('''
                SELECT cl.id, cl.name, cl.section, cl.academic_year,
                       u.username as teacher_name,
                       COUNT(DISTINCT sd.user_id) as student_count
                FROM classrooms cl
                LEFT JOIN users u ON cl.teacher_id = u.id
                LEFT JOIN student_details sd ON sd.classroom_id = cl.id
                WHERE cl.school_id = %s
                GROUP BY cl.id, cl.name, cl.section, cl.academic_year, u.username
                ORDER BY cl.name
            ''', (current_user.school_id,))
        else:
            # Teacher only sees their own classrooms
            cursor.execute('''
                SELECT cl.id, cl.name, cl.section, cl.academic_year,
                       u.username as teacher_name,
                       COUNT(DISTINCT sd.user_id) as student_count
                FROM classrooms cl
                LEFT JOIN users u ON cl.teacher_id = u.id
                LEFT JOIN student_details sd ON sd.classroom_id = cl.id
                WHERE cl.teacher_id = %s AND cl.school_id = %s
                GROUP BY cl.id, cl.name, cl.section, cl.academic_year, u.username
                ORDER BY cl.name
            ''', (current_user.id, current_user.school_id))
        classrooms = [dict(row) for row in cursor.fetchall()]

        # Fetch all teachers for the create/assign form (filtered by school)
        cursor.execute("SELECT id, username FROM users WHERE role = 'teacher' AND school_id = %s ORDER BY username", (current_user.school_id,))
        teachers = [dict(row) for row in cursor.fetchall()]

    return render_template('classrooms/index.html', classrooms=classrooms, teachers=teachers, user=current_user)


# ── Class detail: full student roster ───────────────────────────────────────

@classrooms_bp.route('/classrooms/<int:classroom_id>')
@login_required
def detail(classroom_id):
    db = get_db()
    
    # If student, ensure they can only see their own classroom
    if current_user.role == 'student':
        with db_cursor(db) as cursor:
            cursor.execute('SELECT classroom_id FROM student_details WHERE user_id = %s AND school_id = %s', (current_user.id, current_user.school_id))
            row = cursor.fetchone()
            if not row or row['classroom_id'] != classroom_id:
                flash('Unauthorized access to this classroom.', 'error')
                return redirect(url_for('dashboard.dashboard'))
    else:
        redir = _require_staff()
        if redir:
            return redir
    with db_cursor(db) as cursor:
        cursor.execute('''
            SELECT cl.*, u.username as teacher_name
            FROM classrooms cl
            LEFT JOIN users u ON cl.teacher_id = u.id
            WHERE cl.id = %s AND cl.school_id = %s
        ''', (classroom_id, current_user.school_id))
        classroom = cursor.fetchone()
        if not classroom:
            flash('Classroom not found.', 'error')
            return redirect(url_for('classrooms.index'))

        # Enforce teacher can only see their own classroom (Principals/Admins see all)
        if current_user.role == 'teacher' and classroom['teacher_id'] != current_user.id:
            flash('You do not have access to this classroom.', 'error')
            return redirect(url_for('classrooms.index'))

        cursor.execute('''
            SELECT
                u.id,
                sd.full_name,
                sd.admission_number,
                sd.email,
                sd.gender,
                COALESCE(ROUND(AVG(g.score), 1), 0)  AS avg_grade,
                COUNT(DISTINCT CASE WHEN a.status = %s THEN a.id END) * 100.0 /
                    NULLIF(COUNT(DISTINCT a.id), 0)   AS attendance_pct
            FROM users u
            JOIN student_details sd ON sd.user_id = u.id
            LEFT JOIN grades g ON g.student_id = u.id
            LEFT JOIN attendance a ON a.student_id = u.id
            WHERE sd.classroom_id = %s AND sd.school_id = %s
            GROUP BY u.id, sd.full_name, sd.admission_number, sd.email, sd.gender
            ORDER BY sd.full_name
        ''', ('Present', classroom_id, current_user.school_id))
        students = [dict(row) for row in cursor.fetchall()]

        # Fetch students NOT in this classroom (including those in NO classroom)
        cursor.execute('''
            SELECT u.id, sd.full_name, sd.admission_number, cl.name as current_class
            FROM users u
            JOIN student_details sd ON sd.user_id = u.id
            LEFT JOIN classrooms cl ON sd.classroom_id = cl.id
            WHERE u.role = 'student' AND u.school_id = %s AND (sd.classroom_id IS NULL OR sd.classroom_id != %s)
            ORDER BY sd.full_name
        ''', (current_user.school_id, classroom_id))
        available_students = [dict(row) for row in cursor.fetchall()]

        # For student view: Fetch enrolled courses with their averages
        enrolled_courses = []
        if current_user.role == 'student':
            cursor.execute('''
                SELECT c.*, u.username as teacher_name,
                       COALESCE(ROUND(AVG(g.score), 1), 0) as avg_grade
                FROM courses c 
                JOIN enrollments e ON c.id = e.course_id 
                JOIN users u ON c.teacher_id = u.id
                LEFT JOIN grades g ON (g.course_id = c.id AND g.student_id = %s)
                WHERE e.student_id = %s AND c.school_id = %s
                GROUP BY c.id, u.username
            ''', (current_user.id, current_user.id, current_user.school_id))
            enrolled_courses = [dict(row) for row in cursor.fetchall()]

        # For teacher/admin: Fetch grades matrix
        grades_matrix = {'courses': [], 'students': []}
        if current_user.role in ['admin', 'teacher', 'principal']:
            # 1. Get all courses relevant to this classroom's students
            cursor.execute('''
                SELECT DISTINCT c.id, c.name 
                FROM courses c
                JOIN enrollments e ON c.id = e.course_id
                JOIN student_details sd ON e.student_id = sd.user_id
                WHERE sd.classroom_id = %s AND c.school_id = %s
                ORDER BY c.name
            ''', (classroom_id, current_user.school_id))
            relevant_courses = cursor.fetchall()
            grades_matrix['courses'] = [dict(c) for c in relevant_courses]

            # 2. For each student, get their average grade in each course
            for student in students:
                student_grades = {'full_name': student['full_name'], 'admission_number': student['admission_number'], 'scores': {}}
                for rc in relevant_courses:
                    cursor.execute('''
                        SELECT COALESCE(ROUND(AVG(score), 1), 0) FROM grades 
                        WHERE student_id = %s AND course_id = %s AND school_id = %s
                    ''', (student['id'], rc['id'], current_user.school_id))
                    score = cursor.fetchone()[0]
                    student_grades['scores'][rc['id']] = score
                grades_matrix['students'].append(student_grades)

    return render_template('classrooms/detail.html',
                           classroom=dict(classroom),
                           students=students,
                           available_students=available_students,
                           enrolled_courses=enrolled_courses,
                           grades_matrix=grades_matrix,
                           user=current_user)


# ── Bulk Add Students to Classroom (admin only) ──────────────────────────────

@classrooms_bp.route('/classrooms/<int:classroom_id>/add-students', methods=['POST'])
@login_required
def add_students(classroom_id):
    if current_user.role not in ['teacher', 'admin']:
        flash('Unauthorized access.', 'error')
        return redirect(url_for('classrooms.index'))

    db = get_db()
    # If teacher, verify they own the class
    if current_user.role == 'teacher':
        with db_cursor(db) as cursor:
            cursor.execute('SELECT teacher_id FROM classrooms WHERE id = %s AND school_id = %s', (classroom_id, current_user.school_id))
            cl = cursor.fetchone()
            if not cl or cl['teacher_id'] != current_user.id:
                flash('You can only add students to your own classroom.', 'error')
                return redirect(url_for('classrooms.index'))

    student_ids = request.form.getlist('student_ids')
    if not student_ids:
        flash('No students selected.', 'warning')
        return redirect(url_for('classrooms.detail', classroom_id=classroom_id))

    db = get_db()
    try:
        # Convert IDs to integers for DB safety
        student_int_ids = [int(sid) for sid in student_ids if sid.isdigit()]
        
        if not student_int_ids:
            flash('Invalid student selection.', 'error')
            return redirect(url_for('classrooms.detail', classroom_id=classroom_id))

        with db_cursor(db) as cursor:
            # Loop for SQLite + PostgreSQL compatibility
            for sid in student_int_ids:
                cursor.execute(
                    'UPDATE student_details SET classroom_id = %s WHERE user_id = %s AND school_id = %s',
                    (classroom_id, sid, current_user.school_id)
                )
        db.commit()
        flash(f'Successfully assigned {len(student_int_ids)} student(s) to the class!', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Error assigning students: {str(e)}', 'error')

    return redirect(url_for('classrooms.detail', classroom_id=classroom_id))


# ── QR Attendance System ────────────────────────────────────────────────────

@classrooms_bp.route('/classrooms/<int:classroom_id>/scan')
@login_required
def qr_scanner(classroom_id):
    """Render the live QR scanner page for a classroom."""
    redir = _require_staff()
    if redir: return redir

    db = get_db()
    with db_cursor(db) as cursor:
        cursor.execute("SELECT * FROM classrooms WHERE id = %s AND school_id = %s", (classroom_id, current_user.school_id))
        classroom = cursor.fetchone()
        
        # Fetch courses taught by this teacher so they can select which class they are marking
        if current_user.role == 'admin':
            cursor.execute("SELECT * FROM courses WHERE school_id = %s", (current_user.school_id,))
        else:
            cursor.execute("SELECT * FROM courses WHERE teacher_id = %s AND school_id = %s", (current_user.id, current_user.school_id))
        courses = cursor.fetchall()

    if not classroom:
        flash('Classroom not found.', 'error')
        return redirect(url_for('classrooms.index'))

    return render_template('attendance/scan.html', classroom=classroom, courses=courses, user=current_user)


@classrooms_bp.route('/api/attendance/qr-mark', methods=['POST'])
@login_required
def mark_qr_attendance():
    """API endpoint to mark attendance via scanned QR code."""
    if current_user.role not in ['teacher', 'admin']:
        return {'success': False, 'message': 'Unauthorized'}, 403

    data = request.json
    qr_raw = data.get('qr_code', '') # Layout: "student:<id>:<admission>"
    course_id = data.get('course_id')
    
    if not qr_raw or not course_id:
        return {'success': False, 'message': 'Invalid data'}, 400

    try:
        parts = qr_raw.split(':')
        if len(parts) < 2 or parts[0] != 'student':
            return {'success': False, 'message': 'Invalid QR format'}, 400
        
        student_id = int(parts[1])
        admission_no = parts[2] if len(parts) > 2 else ""
        
        db = get_db()
        from datetime import date
        today = date.today().isoformat()

        with db_cursor(db) as cursor:
            # Verify student exists in the same school
            cursor.execute('SELECT full_name FROM student_details WHERE user_id = %s AND school_id = %s', (student_id, current_user.school_id))
            student = cursor.fetchone()
            if not student:
                return {'success': False, 'message': 'Student not found'}, 404

            # Mark attendance (UPSERT logic: update if already exists for today/student/course)
            # Check if exists
            cursor.execute(
                'SELECT id FROM attendance WHERE student_id = %s AND course_id = %s AND date = %s AND school_id = %s',
                (student_id, course_id, today, current_user.school_id)
            )
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute(
                    'UPDATE attendance SET status = %s WHERE id = %s AND school_id = %s',
                    ('Present', existing['id'], current_user.school_id)
                )
            else:
                cursor.execute(
                    'INSERT INTO attendance (student_id, course_id, date, status, school_id) VALUES (%s, %s, %s, %s, %s)',
                    (student_id, course_id, today, 'Present', current_user.school_id)
                )
        db.commit()
        return {'success': True, 'student_name': student['full_name'], 'admission_no': admission_no}

    except Exception as e:
        return {'success': False, 'message': str(e)}, 500


# ── Export Classroom Data to Excel ──────────────────────────────────────────

@classrooms_bp.route('/classrooms/<int:classroom_id>/export')
@login_required
def export_excel(classroom_id):
    redir = _require_staff()
    if redir:
        return redir

    db = get_db()
    with db_cursor(db) as cursor:
        # 1. Fetch classroom info
        cursor.execute("SELECT name, section, academic_year FROM classrooms WHERE id = %s AND school_id = %s", (classroom_id, current_user.school_id))
        classroom = cursor.fetchone()
        if not classroom:
            flash("Classroom not found", "error")
            return redirect(url_for('classrooms.index'))

        # 2. Fetch relevant courses for this classroom
        cursor.execute('''
            SELECT DISTINCT c.id, c.name 
            FROM courses c
            JOIN enrollments e ON c.id = e.course_id
            JOIN student_details sd ON e.student_id = sd.user_id
            WHERE sd.classroom_id = %s AND c.school_id = %s
            ORDER BY c.name
        ''', (classroom_id, current_user.school_id))
        courses = cursor.fetchall()
        course_list = [dict(c) for c in courses]

        # 3. Fetch students with detailed grades
        cursor.execute('''
            SELECT u.id, sd.full_name, sd.admission_number
            FROM student_details sd
            JOIN users u ON sd.user_id = u.id
            WHERE sd.classroom_id = %s AND sd.school_id = %s
            ORDER BY sd.full_name
        ''', (classroom_id, current_user.school_id))
        students_raw = cursor.fetchall()
        
        student_data = []
        for s in students_raw:
            s_dict = dict(s)
            s_dict['grades'] = {}
            for c in course_list:
                cursor.execute('''
                    SELECT COALESCE(ROUND(AVG(score), 1), 0) FROM grades 
                    WHERE student_id = %s AND course_id = %s AND school_id = %s
                ''', (s['id'], c['id'], current_user.school_id))
                s_dict['grades'][c['id']] = cursor.fetchone()[0]
            student_data.append(s_dict)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 4. Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades Report"

    # Define Styles
    title_font = Font(size=16, bold=True, color="1E1B4B")
    label_font = Font(bold=True, size=10, color="6B7280")
    value_font = Font(bold=True, size=11, color="111827")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Professional Header Section
    ws.merge_cells(f'A1:{get_column_letter(len(course_list) + 3)}1')
    ws['A1'] = "Classroom Academic Intelligence Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    ws['A3'] = "Classroom:"
    ws['A3'].font = label_font
    ws['B3'] = f"{classroom['name']} ({classroom['section'] or 'General'})"
    ws['B3'].font = value_font

    ws['A4'] = "Academic Year:"
    ws['A4'].font = label_font
    ws['B4'] = classroom['academic_year']
    ws['B4'].font = value_font

    # Summary Stats
    total_students = len(student_data)
    ws['F3'] = "Total Students:"
    ws['F3'].font = label_font
    ws['G3'] = total_students
    ws['G3'].font = value_font

    # Main Data Table
    start_row = 7
    
    # Header
    headers = ["Admission No", "Student Name"]
    for c in course_list:
        headers.append(c['name'])
    headers.append("Overall Average")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Data Rows
    for row_num, s in enumerate(student_data, start_row + 1):
        ws.cell(row=row_num, column=1, value=s['admission_number']).border = border
        ws.cell(row=row_num, column=2, value=s['full_name']).border = border
        
        last_course_col = 2
        for i, c in enumerate(course_list):
            val = s['grades'][c['id']]
            cell = ws.cell(row=row_num, column=3+i, value=val)
            cell.border = border
            cell.alignment = center_align
            last_course_col = 3 + i
            
        # Add Excel formula for Average
        avg_col = last_course_col + 1
        start_col_letter = get_column_letter(3)
        end_col_letter = get_column_letter(last_course_col)
        # Using AVERAGE formula in Excel
        formula = f"=IFERROR(ROUND(AVERAGE({start_col_letter}{row_num}:{end_col_letter}{row_num}), 1), 0)"
        avg_cell = ws.cell(row=row_num, column=avg_col, value=formula)
        avg_cell.font = Font(bold=True)
        avg_cell.border = border
        avg_cell.alignment = center_align

    # Column Widths
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    filename = f"Grades_{classroom['name']}_{classroom['academic_year']}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Create classroom (admin only) ────────────────────────────────────────────

@classrooms_bp.route('/classrooms/create', methods=['POST'])
@login_required
def create():
    if current_user.role != 'admin':
        flash('Only admins can create classrooms.', 'error')
        return redirect(url_for('classrooms.index'))

    name = request.form.get('name', '').strip()
    section = request.form.get('section', '').strip()
    teacher_id = request.form.get('teacher_id') or None
    academic_year = request.form.get('academic_year', '2025-2026').strip()

    if not name:
        flash('Classroom name is required.', 'error')
        return redirect(url_for('classrooms.index'))

    db = get_db()
    try:
        with db_cursor(db) as cursor:
            cursor.execute(
                'INSERT INTO classrooms (name, section, teacher_id, academic_year, school_id) VALUES (%s, %s, %s, %s, %s)',
                (name, section, teacher_id, academic_year, current_user.school_id)
            )
        db.commit()
        flash(f'Classroom "{name}" created successfully!', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Error creating classroom: {str(e)}', 'error')

    return redirect(url_for('classrooms.index'))


# ── Assign teacher to classroom (admin only) ─────────────────────────────────

@classrooms_bp.route('/classrooms/<int:classroom_id>/assign', methods=['POST'])
@login_required
def assign_teacher(classroom_id):
    if current_user.role != 'admin':
        flash('Only admins can assign teachers.', 'error')
        return redirect(url_for('classrooms.index'))

    teacher_id = request.form.get('teacher_id') or None
    db = get_db()
    try:
        with db_cursor(db) as cursor:
            cursor.execute(
                'UPDATE classrooms SET teacher_id = %s WHERE id = %s AND school_id = %s',
                (teacher_id, classroom_id, current_user.school_id)
            )
        db.commit()
        flash('Teacher assigned successfully!', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('classrooms.detail', classroom_id=classroom_id))


# ── Delete classroom (admin only) ─────────────────────────────────────────────

@classrooms_bp.route('/classrooms/<int:classroom_id>/delete', methods=['POST'])
@login_required
def delete(classroom_id):
    if current_user.role != 'admin':
        flash('Only admins can delete classrooms.', 'error')
        return redirect(url_for('classrooms.index'))

    db = get_db()
    try:
        with db_cursor(db) as cursor:
            # Unlink students first
            cursor.execute('UPDATE student_details SET classroom_id = NULL WHERE classroom_id = %s AND school_id = %s', (classroom_id, current_user.school_id))
            cursor.execute('DELETE FROM classrooms WHERE id = %s AND school_id = %s', (classroom_id, current_user.school_id))
        db.commit()
        flash('Classroom deleted.', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('classrooms.index'))
